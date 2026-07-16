"""Budget-ID routing for **Other Invoice Types** (hardware / software / subscription).

Routes an invoice to a budget row in `documentation/Budget_ID's.xlsx` ("ID Master" sheet) to recover
its **Budget ID**, **cost center**, **approver**, **offset GL account**, and (as a fallback) its
**supplier number**. NOT used for contractor invoices.

Matching is **vendor-first, then description** (user-specified): narrow to rows whose Supplier (col C)
matches the invoice vendor, then fuzzy-match the invoice's service description(s) against the Product
column (col F). A strong product match returns that exact row; a weak match still returns any value the
vendor's rows unanimously agree on (e.g. every Salesforce row is cost center DT600), so the sheet can
serve as the Copy-Tracker fallback for cost center / approver even when the product line is ambiguous.

Columns (by header, position-independent): ID (B), Supplier (C), Supplier Number (D), Approver (E),
Product (F), DT Cost Center (G), GL ACCOUNT (K).
"""
from __future__ import annotations

import re
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path

from app.config import REPO_ROOT
from app.services.coupa import _normalize_supplier_name

BUDGET_XLSX = REPO_ROOT / "documentation" / "Budget_ID's.xlsx"

_VENDOR_FUZZY_CUTOFF = 88   # supplier (col C) vs invoice vendor
_PRODUCT_MATCH_CUTOFF = 60  # invoice description vs Product (col F, vendor-stripped) to trust a row

# Placeholder values that are not real data (appear in Supplier/Approver/Cost Center/ID cells).
_PLACEHOLDERS = {"", "multiple", "unknown", "none", "non-dt", "open", "n/a", "na", "tbd", "various"}
# A real account/cost-center code: letters+digits ("DT600", "IT802") or bare digits ("679030", "131120").
_CODE_RE = re.compile(r"^\s*([A-Za-z]{1,4}\d{3,6}|\d{3,6})\b")


@dataclass(frozen=True)
class BudgetRow:
    supplier_norm: str
    product_norm: str
    budget_id: str | None
    cost_center: str | None
    approver: str | None
    offset_gl_account: str | None
    supplier_number: str | None


def _code_prefix(value: object) -> str | None:
    """Leading account/cost-center code from a labeled cell ('DT600 - App & Platform' -> 'DT600',
    '679030 - IT SERVICES' -> '679030'); None if the cell isn't a real code ('MULTIPLE', 'OPEN')."""
    s = str(value).strip() if value is not None else ""
    m = _CODE_RE.match(s)
    return m.group(1) if m else None


def _clean_text(value: object) -> str | None:
    """A displayable text value (approver / budget id), or None for blanks and placeholder tokens."""
    s = str(value).strip() if value is not None else ""
    return None if s.lower() in _PLACEHOLDERS else (s or None)


def _norm(value: object) -> str:
    return re.sub(r"\s+", " ", str(value).strip().lower()) if value is not None else ""


def _load_budget_rows(path: Path) -> list[BudgetRow]:
    rows: list[BudgetRow] = []
    if not path.exists():
        return rows
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return rows
    try:
        ws = wb[wb.sheetnames[0]]
        it = ws.iter_rows(values_only=True)
        header = next(it, None)
        if not header:
            return rows
        cols = {str(h).strip(): i for i, h in enumerate(header) if h is not None}
        ci_sup, ci_supno = cols.get("Supplier"), cols.get("Supplier Number")
        ci_id, ci_appr = cols.get("ID"), cols.get("Approver")
        ci_prod, ci_cc, ci_gl = cols.get("Product"), cols.get("DT Cost Center"), cols.get("GL ACCOUNT")
        if ci_sup is None or ci_prod is None:
            return rows

        def cell(row, i):
            return row[i] if i is not None and i < len(row) else None

        for row in it:
            supplier_norm = _normalize_supplier_name(str(cell(row, ci_sup) or ""))
            if not supplier_norm or supplier_norm in _PLACEHOLDERS:
                continue  # skip "MULTIPLE" / blank vendors — they can't be routed to
            supno = _clean_text(cell(row, ci_supno))
            rows.append(BudgetRow(
                supplier_norm=supplier_norm,
                product_norm=_norm(cell(row, ci_prod)),
                budget_id=_clean_text(cell(row, ci_id)),
                cost_center=_code_prefix(cell(row, ci_cc)),
                approver=_clean_text(cell(row, ci_appr)),
                offset_gl_account=_code_prefix(cell(row, ci_gl)),
                # Supplier numbers are bare digits here; canonical form elsewhere is "RAC-<digits>".
                supplier_number=(f"RAC-{supno}" if supno and supno.isdigit() else supno),
            ))
    finally:
        wb.close()
    return rows


@lru_cache(maxsize=1)
def _budget_rows() -> list[BudgetRow]:
    return _load_budget_rows(BUDGET_XLSX)


def _vendor_candidates(vendor_norm: str, rows: list[BudgetRow]) -> list[BudgetRow]:
    """Budget rows whose Supplier matches the invoice vendor: exact-normalized, else high fuzzy."""
    if not vendor_norm:
        return []
    exact = [r for r in rows if r.supplier_norm == vendor_norm]
    if exact:
        return exact
    from rapidfuzz import fuzz

    return [r for r in rows if fuzz.token_sort_ratio(vendor_norm, r.supplier_norm) >= _VENDOR_FUZZY_CUTOFF]


def _consensus(values: list[str | None]) -> str | None:
    """The single distinct non-empty value shared by all candidates, else None (ambiguous)."""
    distinct = {v for v in values if v}
    return next(iter(distinct)) if len(distinct) == 1 else None


def budget_route(vendor_name: str | None, descriptions: list[str]) -> dict[str, str | None]:
    """Route an Other invoice to its budget row. Vendor-first, then service description -> Product (F).

    Returns budget_id / cost_center / approver / offset_gl_account / supplier_number. A confident
    product match yields that row's values; a weak match still yields any field the vendor's rows all
    agree on (so cost center / supplier number survive an ambiguous product line).
    """
    empty = {"budget_id": None, "cost_center": None, "approver": None,
             "offset_gl_account": None, "supplier_number": None}
    vendor_norm = _normalize_supplier_name(vendor_name or "")
    candidates = _vendor_candidates(vendor_norm, _budget_rows())
    if not candidates:
        return empty

    from rapidfuzz import fuzz

    # Strip the vendor's own tokens from the Product cell before matching: rows read "Salesforce - B2B"
    # but the invoice line is just "B2B Commerce", and we've already filtered to this vendor.
    vendor_tokens = set(vendor_norm.split())
    desc = " ".join(d for d in descriptions if d).strip().lower()
    best_row, best_score = None, -1.0
    if desc:
        for r in candidates:
            product = " ".join(t for t in r.product_norm.split() if t not in vendor_tokens)
            if not product:
                continue
            score = fuzz.token_set_ratio(desc, product)
            if score > best_score:
                best_row, best_score = r, score

    strong = best_row is not None and best_score >= _PRODUCT_MATCH_CUTOFF
    primary = best_row if strong else None
    # Field value: the matched row's value if we trust the match, else the vendor-wide consensus.
    def resolve(attr: str) -> str | None:
        if primary is not None and getattr(primary, attr):
            return getattr(primary, attr)
        return _consensus([getattr(r, attr) for r in candidates])

    return {
        "budget_id": primary.budget_id if primary else _consensus([r.budget_id for r in candidates]),
        "cost_center": resolve("cost_center"),
        "approver": resolve("approver"),
        "offset_gl_account": resolve("offset_gl_account"),
        "supplier_number": resolve("supplier_number"),
    }
