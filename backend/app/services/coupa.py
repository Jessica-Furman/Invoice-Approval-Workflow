"""Coupa import-CSV generation for matched contractor invoices.

Turns a matched `Invoice` (+ its line items + linked Clarity data) into a Coupa-ready import CSV that
matches **`documentation/Coupa sample invoice.xlsx`** exactly — the same column set, header names and
row layout Coupa ingests. Any column we can't source from the invoice/Clarity is left **blank** (no
placeholder tokens): a human fills those in the downloaded file if needed.

Structure of the emitted file (exactly mirrors the sample workbook):
    row 1  -> the `Invoice` record-type column schema     (177 columns)
    row 2  -> the `Invoice Line` record-type column schema (73 columns)
    row 3  -> the invoice's `Invoice` header data row
    row 4+ -> one `Invoice Line` data row per contractor line item

Each data row is tagged in its first cell by record type ("Invoice" / "Invoice Line"), positionally
aligned to that type's schema row. Only the fields we have are populated:

    Header : Invoice Number*, Supplier Name/Number, Invoice Date*, Submit For Approval?,
             Line Level Taxation*, Requester Name (the approver, from the copy tracker),
             Chart of Accounts* (matched only), Currency. Status is intentionally BLANK.
    Line   : Line Number, Description*, Price* (rate), Quantity (hours), Unit of Measure* (HOUR),
             Account Segment 1 (company code), 2 (cost center), 3 (CapEx/OpEx GL code).
             PO / Requester / CapEx-OpEx-label columns exist in the schema but are left blank.
"""
from __future__ import annotations

import csv
import io
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import date
from functools import lru_cache
from pathlib import Path

from sqlalchemy import select
from sqlalchemy.orm import Session

from app import models
from app.config import REPO_ROOT
from app.services.matching import _line_period, countable_status_filter

# --- Column schemas — the EXACT Coupa template (documentation/Coupa sample invoice.xlsx) ----------
# Two record types share the file, distinguished by the first cell. These lists are the sample
# workbook's row 1 ("Invoice") and row 2 ("Invoice Line") headers, verbatim (asterisks and all).

INVOICE_HEADER_COLUMNS = [
    'Invoice', 'Invoice Number*', 'Supplier Name',
    'Supplier Number', 'Status', 'Invoice Date*',
    'Submit For Approval?', 'Handling Amount', 'Misc Amount',
    'Shipping Amount', 'Line Level Taxation*', 'Tax Amount',
    'Tax Rate', 'Tax Code', 'Tax Rate Type',
    'Supplier Note', 'Payment Terms', 'Shipping Terms',
    'Requester Email', 'Requester Name', 'Requester Lookup Name',
    'Chart of Accounts*', 'Currency', 'Contract Number',
    'Image Scan Filename', 'Image Scan URL', 'Local Currency Net',
    'Taxes In Origin Country Currency', 'Local Currency Gross', 'Delivery Number',
    'Delivery Date', 'Margin Scheme', 'Invoice Issuance Time',
    'Cash Register Operator', 'Means Of Payment', 'Unique Identification Code Of Cash Receipt',
    'Security Code Of Issuer', 'Cash Accounting Scheme Reference', 'Exchange Rate',
    'Gross Total', 'Invoice Control Total', 'Late Payment Penalties',
    'Credit Reason', 'Early Payment Provisions', 'Pre-Payment Date',
    'Self Billing Reference', 'Discount Amount', 'Reverse Charge Reference',
    'Discount %', 'Credit Note differences with Original Invoice', 'Original Value of Supply',
    'Correct Value of Supply', 'Customs Declaration Number', 'Customs Office',
    'Customs Declaration Date', 'Payment Order Reference', 'Payment Order Number',
    'Resolution Number', 'Amount of advance payment received', 'Type of Relationship',
    'Invoice Reference Number (IRN)', 'Series', 'Folio Number',
    'Buyer Representative Name', 'Supplier Representative Name', 'Payment Schedule Terms',
    'Transaction UUID', 'Transaction Notification Date', 'VAT Chargeability System',
    'Ship To Name', 'Ship To Id', 'Ship To Attention',
    'Ship To Street1', 'Ship To Street2', 'Ship To City',
    'Ship To State', 'Ship To Postal Code', 'Ship to Country Code',
    'Ship to Country Name', 'Ship to Location Code', 'Ship to VAT ID',
    'Ship to Local Tax Number', 'Bill To Address Id', 'Bill To Address Legal Entity Name',
    'Bill To Address Street', 'Bill To Address City', 'Bill To Address Postal Code',
    'Bill To Address Country Code', 'Bill To Address Location Code', 'Bill To Address VAT ID',
    'Bill To Address Local Tax Number', 'Override Tax Country Code', 'Remit To Address Street1',
    'Remit To Address Street2', 'Remit To Address City', 'Remit To Address State',
    'Remit To Address Postal Code', 'Remit To Address Country Code', 'Remit To Code',
    'Remit To Tax Prefix', 'Remit To Tax Number', 'Remit To Tax Country Code',
    'Remit To VAT ID', 'Remit To Local Tax Number', 'Invoice From Address Street1',
    'Invoice From Address Street2', 'Invoice From Address City', 'Invoice From Address State',
    'Invoice From Address Postal Code', 'Invoice From Address Country Code', 'Invoice From Code',
    'Ship From Address Street1', 'Ship From Address Street2', 'Ship From Address City',
    'Ship From Address State', 'Ship From Address Postal Code', 'Ship From Address Country Code',
    'Ship From Code', 'Original invoice number', 'Original invoice date',
    'Is Credit Note', 'Disputed Invoice Number', 'Dispute Resolution Credit Note Number',
    'Supplier Tax Number', 'Buyer Tax Number', 'Date of Discovery of Facts Decisive for Correction',
    'Place Of Supply', 'Split Payment Mechanism', 'Endorsement On Invoices',
    'New Means Of Transport', 'Place Of Issuance', 'Amount Of Advance Payment',
    'Supplier Invoice Issuer Name', 'Supplier Invoice Reviewer Name', 'Supplier Payment Collector Name',
    'Signed QR Code', 'State Tax ID Number', 'State Tax ID Number for Substitute Taxpayer',
    'Municipality Tax ID Number', 'Serial Code of Fiscal Invoice', 'Verification Code',
    'Type of Document', 'Protocol Number', 'Nature of Operation',
    'Type of Operation', 'Freight Type', 'Vehicle License Plate',
    'National Enrollment of Conveyor', 'Volume Amount', 'Volume Gross Weight',
    'Volume Liquid Weight', 'Volume Brand', 'Volume Type',
    'Volume Numbering', 'Payment Agreement Notes', 'Attachment 1',
    'Attachment 2', 'Attachment 3', 'Attachment 4',
    'Attachment 5', 'Attachment 6', 'Attachment 7',
    'Attachment 8', 'Attachment 9', 'Attachment 10',
    'Ship To Street3', 'Ship To Street4', 'Remit To Address Street3',
    'Remit To Address Street4', 'Invoice From Address Street3', 'Invoice From Address Street4',
    'Ship From Address Street3', 'Ship From Address Street4', 'Invoice Group',
    'Batch ID', 'Posting Date', 'Comments to Supplier',
]

INVOICE_LINE_COLUMNS = [
    'Invoice Line', 'Invoice Number*', 'Supplier Name',
    'Supplier Number', 'Line Number', 'Description*',
    'Supplier Part Number', 'Auxiliary Part Number', 'Price*',
    'Quantity', 'Bulk Price', 'Bulk Price Qty',
    'Bulk Price UOM', 'Bulk Price Conversion Numerator', 'Bulk Price Conversion Denominator',
    'Unit of Measure*', 'Category', 'Subcategory',
    'Deductibility', 'PO Number', 'PO Line Number',
    'Account Name', 'Account Code', 'Billing Notes',
    'Account Segment 1', 'Account Segment 2', 'Account Segment 3',
    'Account Segment 4', 'Account Segment 5', 'Account Segment 6',
    'Account Segment 7', 'Account Segment 8', 'Account Segment 9',
    'Account Segment 10', 'Account Segment 11', 'Account Segment 12',
    'Account Segment 13', 'Account Segment 14', 'Account Segment 15',
    'Account Segment 16', 'Account Segment 17', 'Account Segment 18',
    'Account Segment 19', 'Account Segment 20', 'Budget Period Name',
    'Net Weight', 'Weight UOM', 'Price per Weight',
    'Match Reference', 'Delivery Note Number', 'Original Date Of Supply',
    'Commodity Name', 'HSN/SAC', 'UNSPSC',
    'Adjustment Type', 'Fiscal Code', 'Classification of Goods',
    'Municipality Taxation Code', 'Item Barcode', 'Line Tax Amount',
    'Line Tax Rate', 'Line Tax Code', 'Line Tax Rate Type',
    'Exempt from Tax?', 'Reverse Charge?', 'Out of Scope?',
    'Line Tax Location', 'Line Tax Description', 'Line Tax Supply Date',
    'Line Nature of Tax', 'Line Product Tax Classification', 'Customer Accounting?',
    'Exclude from tax',
]

# --- Accounting segment mappings (documentation/csv_rules) ---------------------------------------
# Account Segment 1 = company code, Segment 2 = cost center, Segment 3 = the CapEx/OpEx GL code.
# (We emit only the CODE, never a CapEx/OpEx label.)
COMPANY_CODE = {"RAC": "5", "ACIMA": "67"}
# Company cost center — used for CAPEX work (and as the CapEx/OpEx-agnostic default). OPEX work is
# coded to the contractor's Clarity department DT code instead (see `cost_center_for`).
COST_CENTER = {"RAC": "H0003", "ACIMA": "AC000"}
# CapEx/OpEx GL code. CAPEX is uniform; OPEX is 667040 for almost every vendor, but 667070 for a few
# (Cigniti, Tata Consultancy) — matched case-insensitively against the invoice vendor name.
CAPEX_GL_CODE = "246010"
OPEX_GL_CODE_DEFAULT = "667040"
_OPEX_GL_VENDOR_OVERRIDES = {"cigniti": "667070", "tata": "667070"}
# Kept for callers that only need the set of valid CapEx/OpEx labels (e.g. `_norm_capex`) and for the
# "other invoice" GL->label reverse map in reporting.py; contractor GL codes go through `gl_code_for`.
CAPEX_OPEX_CODE = {"CAPEX": CAPEX_GL_CODE, "OPEX": OPEX_GL_CODE_DEFAULT}


def gl_code_for(capex_opex: str | None, vendor_name: str | None) -> str:
    """The CapEx/OpEx GL code for a line: CAPEX -> 246010; OPEX -> 667070 for override vendors
    (Cigniti/Tata), else 667040. Empty string when the CapEx/OpEx class is unknown."""
    cx = _norm_capex(capex_opex)
    if cx == "CAPEX":
        return CAPEX_GL_CODE
    if cx == "OPEX":
        v = (vendor_name or "").lower()
        for key, code in _OPEX_GL_VENDOR_OVERRIDES.items():
            if key in v:
                return code
        return OPEX_GL_CODE_DEFAULT
    return ""


def cost_center_for(company: str | None, capex_opex: str | None, dt_code: str | None) -> str:
    """The cost center for a line: OPEX -> the contractor's Clarity department DT code (e.g. DT500);
    CAPEX (or unknown class) -> the company cost center (RAC H0003 / ACIMA AC000). Blank if neither
    resolves."""
    if _norm_capex(capex_opex) == "OPEX":
        return dt_code or ""
    return COST_CENTER.get(company, "")


def project_cost_center_display(
    company: str | None, capex_opex: str | None, dt_codes: set[str] | None
) -> str:
    """Cost center shown for a PROJECT row (drawer / Excel), which can span multiple contractors:
    OPEX -> the distinct DT code(s) of the contractors who worked it (joined if they differ);
    CAPEX (or unknown) -> the company cost center. Blank if unresolved."""
    if _norm_capex(capex_opex) == "OPEX":
        return ", ".join(sorted(c for c in (dt_codes or set()) if c))
    return COST_CENTER.get(company, "")

# Company is resolved from the project's Line of Business in documentation/Company_by_Project.csv,
# keyed by Project ID. LOB "/Upbound" or "/Upbound/RAC" -> RAC; "/Upbound/Acima" -> ACIMA.
COMPANY_BY_PROJECT_CSV = REPO_ROOT / "documentation" / "Company_by_Project.csv"
_LOB_RAC = {"/upbound", "/upbound/rac"}
_LOB_ACIMA = {"/upbound/acima"}

# Fallback when a project ID isn't in the mapping: infer from the project name.
# \b avoids false hits like "tRACking"; "RACPad"/"RAConnect"/"RAC ..." all match \bRAC.
_RAC_RE = re.compile(r"\brac", re.IGNORECASE)
_ACIMA_RE = re.compile(r"\bacima", re.IGNORECASE)


def _load_project_company_map(path: Path) -> dict[str, str]:
    """Read Company_by_Project.csv into {Project ID -> 'RAC' | 'ACIMA'} via its Line of Business."""
    mapping: dict[str, str] = {}
    if not path.exists():
        return mapping
    with open(path, encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            pid = (row.get("Project ID") or "").strip()
            lob = (row.get("Line of Business") or "").strip().lower()
            if not pid:
                continue
            if lob in _LOB_ACIMA:
                mapping[pid] = "ACIMA"
            elif lob in _LOB_RAC:
                mapping[pid] = "RAC"
    return mapping


@lru_cache(maxsize=1)
def _project_company_map() -> dict[str, str]:
    """Cached Project ID -> company lookup (loaded once from COMPANY_BY_PROJECT_CSV)."""
    return _load_project_company_map(COMPANY_BY_PROJECT_CSV)


def _load_project_lob_map(path: Path) -> dict[str, str]:
    """Read Company_by_Project.csv into {Project ID -> raw Line of Business string}."""
    mapping: dict[str, str] = {}
    if not path.exists():
        return mapping
    with open(path, encoding="utf-8-sig", newline="") as f:
        for row in csv.DictReader(f):
            pid = (row.get("Project ID") or "").strip()
            lob = (row.get("Line of Business") or "").strip()
            if pid and lob:
                mapping[pid] = lob
    return mapping


@lru_cache(maxsize=1)
def _project_lob_map() -> dict[str, str]:
    """Cached Project ID -> raw LOB string (e.g. '/Upbound/Acima')."""
    return _load_project_lob_map(COMPANY_BY_PROJECT_CSV)


def project_accounting(project_id: str | None, investment_name: str | None = None) -> dict[str, str | None]:
    """Accounting attributes for a project: its company, cost center, and raw Line of Business.

    Used to enrich the Excel Projects sheet. Company/cost center come from the same RAC/ACIMA rules the
    Coupa CSV uses; LOB is the raw string from Company_by_Project.csv.
    """
    company = _company_for_project(project_id, investment_name)
    return {
        "company": company,
        "cost_code": COMPANY_CODE.get(company),   # company code: RAC -> 5, ACIMA -> 67
        "cost_center": COST_CENTER.get(company),
        "lob": _project_lob_map().get(project_id) if project_id else None,
    }


# --- Supplier Number lookup (documentation/Active supplier list.csv) -----------------------------
# The vendor->supplier-number mapping. Supplier "Name" is messy: a trailing "(zip)" and sometimes a
# "-<digits>" suffix (the supplier number), e.g. "AVASOFT INC-317598 (08820)" -> RAC-317598. We strip
# those, normalize, and match the invoice vendor exact-first, then fuzzy.
ACTIVE_SUPPLIER_CSV = REPO_ROOT / "documentation" / "Active supplier list.csv"
_SUPPLIER_FUZZY_CUTOFF = 88  # token_sort_ratio; high so we don't grab a similarly-named different vendor


# Legal-entity suffixes stripped (as a trailing run, with spaces removed) to compare a vendor to a
# supplier by core name — handles "… Limited" vs "… Inc" and spaceless parses ("CIGNITITECHNOLOGIESLIMITED").
_DESPACE_SUFFIX_RE = re.compile(r"(incorporated|corporation|limited|company|llc|ltd|inc|corp|llp|plc|pvt|gmbh)+$")


def _normalize_supplier_name(name: str | None) -> str:
    """Canonical key for matching a vendor to a supplier 'Name': drop trailing '(...)' and '-<digits>'
    suffixes, lowercase, and reduce punctuation to single spaces."""
    s = re.sub(r"\s*\([^)]*\)\s*$", "", name or "")  # trailing "(zip/code)"
    s = re.sub(r"-\d+\s*$", "", s)                    # trailing "-<supplier digits>"
    s = re.sub(r"[^a-z0-9]+", " ", s.lower())
    s = re.sub(r"\bcom\b", " ", s)                    # drop a ".com" domain token ("salesforce com inc")
    return re.sub(r"\s+", " ", s).strip()


def _despace_key(norm: str) -> str:
    """A loose key for unambiguous core-name matching: remove spaces and a trailing legal suffix, so
    'cigniti technologies inc' and a spaceless 'cignititechnologieslimited' both reduce to the same key."""
    return _DESPACE_SUFFIX_RE.sub("", norm.replace(" ", ""))


def _load_supplier_index(path: Path) -> tuple[dict[str, str], list[str], list[str], dict[str, set[str]]]:
    """Read the supplier list into: exact {norm name -> number}, parallel norm-name + number lists, and
    {despaced core key -> set of numbers} (used only when a key maps to exactly one supplier)."""
    exact: dict[str, str] = {}
    names: list[str] = []
    numbers: list[str] = []
    despaced: dict[str, set[str]] = defaultdict(set)
    if not path.exists():
        return exact, names, numbers, {}
    with open(path, encoding="utf-8-sig", errors="ignore", newline="") as f:
        reader = csv.reader(f)
        header = next(reader, None)
        if not header:
            return exact, names, numbers, {}
        try:
            ni, si = header.index("Name"), header.index("Supplier Number")
        except ValueError:
            return exact, names, numbers, {}
        for row in reader:
            if len(row) <= max(ni, si):
                continue
            number = row[si].strip()
            norm = _normalize_supplier_name(row[ni])
            if not number or not norm:
                continue
            exact.setdefault(norm, number)  # first occurrence wins
            names.append(norm)
            numbers.append(number)
            key = _despace_key(norm)
            if key:
                despaced[key].add(number)
    return exact, names, numbers, dict(despaced)


@lru_cache(maxsize=1)
def _supplier_index() -> tuple[dict[str, str], list[str], list[str], dict[str, set[str]]]:
    return _load_supplier_index(ACTIVE_SUPPLIER_CSV)


@lru_cache(maxsize=512)
def supplier_number_for(vendor_name: str | None) -> str | None:
    """The Coupa Supplier Number for an invoice vendor, or None if not confidently found.

    1) exact normalized match; 2) unambiguous core-name match (spaceless, legal suffix dropped — handles
    "Limited" vs "Inc" and spaceless parses); 3) high-confidence fuzzy (token_sort_ratio)."""
    if not vendor_name:
        return None
    exact, names, numbers, despaced = _supplier_index()
    norm = _normalize_supplier_name(vendor_name)
    if not norm:
        return None
    if norm in exact:
        return exact[norm]
    key = _despace_key(norm)
    cands = despaced.get(key) if key else None
    if cands and len(cands) == 1:  # exactly one supplier shares this core name
        return next(iter(cands))
    from rapidfuzz import fuzz, process

    hit = process.extractOne(norm, names, scorer=fuzz.token_sort_ratio, score_cutoff=_SUPPLIER_FUZZY_CUTOFF)
    return numbers[hit[2]] if hit else None


# --- Approver lookup (documentation/2026_Copy Tracker.xlsx) --------------------------------------
# The copy tracker is the running log of every invoice already filed: one row per allocation line, with
# the invoice's Vendor, Invoice Number, and the person who "Approved By" it. When we generate a Coupa CSV
# we look the approver up by (vendor, invoice number) and drop the name into the sample's "<<Approver
# Name>>" placeholder — the "Requester Name" column. Invoice number is the strong key (vendor names vary
# between the tracker and a parsed invoice), disambiguated by vendor only when a number is shared.
COPY_TRACKER_XLSX = REPO_ROOT / "documentation" / "2026_Copy Tracker.xlsx"
_APPROVER_VENDOR_CUTOFF = 80  # token_sort_ratio for picking among same-invoice-number vendor variants


def _norm_invno(value: object) -> str:
    """Canonical key for an invoice number across the tracker (may be int/float) and our text: as a
    string, integer-valued floats de-pointed (6347.0 -> '6347'), lowercased, inner whitespace dropped."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return re.sub(r"\s+", "", str(value).strip().lower())


def _load_approver_index(path: Path) -> dict[str, list[tuple[str, str]]]:
    """Read the copy tracker into {normalized invoice number -> [(normalized vendor, approver), ...]}.

    Only the Vendor, Invoice Number, and Approved By columns are read. Rows without an invoice number or
    an approver are skipped. Duplicate (vendor, approver) pairs per number are collapsed.
    """
    index: dict[str, list[tuple[str, str]]] = {}
    if not path.exists():
        return index
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return index
    try:
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            return index
        cols = {str(h).strip(): i for i, h in enumerate(header) if h is not None}
        vi = cols.get("Vendor")
        # Header cell is literally "Invoice Number"; "Approved By" holds the approver.
        ii = cols.get("Invoice Number")
        ai = cols.get("Approved By")
        if ii is None or ai is None:
            return index
        for row in rows:
            if ii >= len(row) or ai >= len(row):
                continue
            invno = _norm_invno(row[ii])
            approver = (str(row[ai]).strip() if row[ai] is not None else "")
            if not invno or not approver:
                continue
            vendor = _normalize_supplier_name(str(row[vi])) if vi is not None and vi < len(row) else ""
            bucket = index.setdefault(invno, [])
            pair = (vendor, approver)
            if pair not in bucket:
                bucket.append(pair)
    finally:
        wb.close()
    return index


@lru_cache(maxsize=1)
def _approver_index() -> dict[str, list[tuple[str, str]]]:
    return _load_approver_index(COPY_TRACKER_XLSX)


@lru_cache(maxsize=512)
def approver_for(vendor_name: str | None, invoice_number: str | None) -> str | None:
    """The approver ("Approved By") for an invoice from the copy tracker, or None if not found.

    Matched by invoice number first. If several tracker rows share that number, the vendor picks between
    them: exact normalized vendor match, else a high-confidence fuzzy match; if the vendor can't be told
    apart but every candidate shares one approver, that approver is returned.
    """
    invno = _norm_invno(invoice_number)
    if not invno:
        return None
    candidates = _approver_index().get(invno)
    if not candidates:
        return None
    approvers = {a for _, a in candidates}
    if len(approvers) == 1:  # unambiguous regardless of vendor
        return next(iter(approvers))

    norm_vendor = _normalize_supplier_name(vendor_name)
    if norm_vendor:
        for vendor, approver in candidates:
            if vendor == norm_vendor:
                return approver
        from rapidfuzz import fuzz, process

        vendors = [v for v, _ in candidates]
        hit = process.extractOne(
            norm_vendor, vendors, scorer=fuzz.token_sort_ratio, score_cutoff=_APPROVER_VENDOR_CUTOFF
        )
        if hit:
            return candidates[hit[2]][1]
    return None


# --- Tracker accounting lookup (documentation/2026_Copy Tracker.xlsx) ----------------------------
# The copy tracker also records, per filed invoice, its Cost Center (column G) and the Offset GL
# Account (column F) — the GL to offset, i.e. the CapEx/OpEx code. We route an OTHER invoice by its
# invoice number to these two values for display. An invoice may span several allocation rows with
# different values; we return every distinct value (comma-joined, first-seen order).
_TRACKER_COST_CENTER_COL = "Cost Center"     # column G
_TRACKER_OFFSET_GL_COL = "Offset GL Account"  # column F (= CapEx/OpEx GL code)


def _load_tracker_accounting_index(path: Path) -> dict[str, list[tuple[str, str]]]:
    """{normalized invoice number -> [(cost center, offset GL account), ...]} from the copy tracker."""
    index: dict[str, list[tuple[str, str]]] = {}
    if not path.exists():
        return index
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return index
    try:
        ws = wb[wb.sheetnames[0]]
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        if not header:
            return index
        cols = {str(h).strip(): i for i, h in enumerate(header) if h is not None}
        ii = cols.get("Invoice Number")
        cci = cols.get(_TRACKER_COST_CENTER_COL)
        gli = cols.get(_TRACKER_OFFSET_GL_COL)
        if ii is None:
            return index
        for row in rows:
            if ii >= len(row):
                continue
            invno = _norm_invno(row[ii])
            if not invno:
                continue
            def _cell(i: int | None) -> str:
                return str(row[i]).strip() if i is not None and i < len(row) and row[i] is not None else ""
            index.setdefault(invno, []).append((_cell(cci), _cell(gli)))
    finally:
        wb.close()
    return index


@lru_cache(maxsize=1)
def _tracker_accounting_index() -> dict[str, list[tuple[str, str]]]:
    return _load_tracker_accounting_index(COPY_TRACKER_XLSX)


def _distinct_join(values: list[str]) -> str | None:
    """Distinct non-empty values in first-seen order, comma-joined (None if there are none)."""
    seen: list[str] = []
    for v in values:
        if v and v not in seen:
            seen.append(v)
    return ", ".join(seen) if seen else None


@lru_cache(maxsize=512)
def tracking_accounting_for(invoice_number: str | None) -> dict[str, str | None]:
    """Cost center + offset GL account for an invoice, routed by invoice number in the copy tracker.

    Returns {'cost_center', 'offset_gl_account'} with every distinct value across the invoice's
    allocation rows (comma-joined), or None per field when the tracker has nothing.
    """
    invno = _norm_invno(invoice_number)
    rows = _tracker_accounting_index().get(invno, []) if invno else []
    return {
        "cost_center": _distinct_join([cc for cc, _ in rows]),
        "offset_gl_account": _distinct_join([gl for _, gl in rows]),
    }


# --- Fixed values (documentation/csv_rules) ------------------------------------------------------
CURRENCY_DEFAULT = "USD"
UNIT_OF_MEASURE = "EA"  # Unit of Measure* — matches documentation/Coupa sample invoice.xlsx ("EA")
LINE_LEVEL_TAXATION = "no"
SUBMIT_FOR_APPROVAL = "No"          # per csv_rules: default to No until the workflow is trusted
# Chart of Accounts value — the literal name of the Coupa Chart of Accounts, exactly as in
# documentation/Coupa sample invoice.xlsx. Used for a fully-matched invoice; drafts/flagged leave it blank.
CHART_OF_ACCOUNTS_MATCHED = "Chart of Accounts"


def _fmt_date(d: date | None) -> str:
    """Coupa dates are MM/DD/YYYY."""
    return d.strftime("%m/%d/%Y") if d else ""


def _fmt_money(n: float | None) -> str:
    return f"{n:.2f}" if n is not None else ""


def _fmt_num(n: float | None) -> str:
    """Hours: drop a pointless trailing .0 (40.0 -> 40) but keep real decimals (37.5)."""
    if n is None:
        return ""
    return str(int(n)) if float(n).is_integer() else str(n)


def _project_for(li: models.InvoiceLineItem, db: Session | None) -> models.ClarityProject | None:
    """Look up the Clarity project behind a line's matched timesheet, if we have one."""
    ts = li.matched_clarity
    if ts is None or not ts.project_id or db is None:
        return None
    return db.scalars(
        select(models.ClarityProject).where(models.ClarityProject.project_id == ts.project_id)
    ).first()


def _company_for_project(project_id: str | None, investment_name: str | None) -> str | None:
    """Resolve a project's company ('RAC' | 'ACIMA').

    Primary: the project's Line of Business in Company_by_Project.csv, keyed by Project ID.
    Fallback: infer from the project name. None when neither resolves (caller leaves segments blank).
    """
    if project_id:
        company = _project_company_map().get(project_id)
        if company:
            return company
    name = investment_name or ""
    if _ACIMA_RE.search(name):
        return "ACIMA"
    if _RAC_RE.search(name):
        return "RAC"
    return None


def _norm_capex(val: str | None) -> str | None:
    """Normalize a CapEx/OpEx value to 'CAPEX' | 'OPEX', else None."""
    if val:
        v = val.strip().upper()
        if v in CAPEX_OPEX_CODE:
            return v
    return None


def _company_for(li: models.InvoiceLineItem) -> str | None:
    """The company for a line, from its matched Clarity timesheet's project (see _company_for_project)."""
    ts = li.matched_clarity
    if ts is None:
        return None
    return _company_for_project(ts.project_id, ts.investment_name)


def _capex_opex_for(
    li: models.InvoiceLineItem, project: models.ClarityProject | None
) -> str | None:
    """The line's CapEx/OpEx classification ('CAPEX' | 'OPEX'), from Clarity, else None."""
    ts = li.matched_clarity
    return _norm_capex((ts.capex_opex if ts else None) or (project.capex_opex if project else None))


def _segment_values(company: str | None, cost_center: str, gl_code: str) -> dict[str, str]:
    """The three Account Segment values; blank where unresolved. Segment 1 = company code,
    Segment 2 = cost center (already resolved — CAPEX company code / OPEX DT code), Segment 3 = the
    CapEx/OpEx GL code (already resolved for the vendor). No CapEx/OpEx label is ever written."""
    return {
        "Account Segment 1": COMPANY_CODE.get(company, ""),
        "Account Segment 2": cost_center,
        "Account Segment 3": gl_code,
    }


def _clarity_project_breakdown(
    li: models.InvoiceLineItem, inv: models.Invoice, db: Session | None
) -> dict[str, float] | None:
    """Sum a matched contractor's in-period Clarity hours grouped by project (Investment Name).

    Same billable filter as matching — counted (Posted or Submitted), not time-off, Date
    Worked within the line's period. A contractor billed on this line may have logged hours against
    more than one Clarity investment in the period; the caller uses each project's share of these
    hours to split the line's dollar amount proportionally. Returns None when it can't be computed
    (no db / no matched Clarity row / zero in-period hours), so the caller attributes the line's full
    amount to "Unresolved" instead.
    """
    ts = li.matched_clarity
    if db is None or ts is None or not ts.contractor_name_normalized:
        return None
    start, end = _line_period(li, inv)
    q = select(models.ClarityTimesheet).where(
        models.ClarityTimesheet.contractor_name_normalized == ts.contractor_name_normalized,
        countable_status_filter(),  # Posted OR Submitted
        models.ClarityTimesheet.is_time_off.is_(False),
    )
    if start and end:
        q = q.where(
            models.ClarityTimesheet.date_worked >= start,
            models.ClarityTimesheet.date_worked <= end,
        )
    groups: dict[str, float] = {}
    for e in db.scalars(q).all():
        label = e.investment_name or e.project_id or "Unresolved"
        groups[label] = groups.get(label, 0.0) + (e.hours or 0.0)
    groups = {k: round(v, 2) for k, v in groups.items() if v}
    return groups or None


@dataclass
class ProjectSplit:
    """One (contractor, project) slice of a contractor's in-period Clarity hours, with the accounting
    attributes needed to code its own Coupa CSV line."""

    project_id: str | None
    project_name: str | None
    hours: float
    company: str | None       # 'RAC' | 'ACIMA' | None
    capex_opex: str | None    # 'CAPEX' | 'OPEX' | None


def _clarity_project_splits(
    li: models.InvoiceLineItem, inv: models.Invoice, db: Session | None
) -> list[ProjectSplit] | None:
    """A matched contractor's in-period Clarity hours split per project (Investment), each carrying
    its company + CapEx/OpEx so it can be coded on its own CSV line.

    Same billable filter as the other breakdowns (Posted/Submitted, not time-off, Date Worked within
    the line's period). Groups by project id; per group sums hours, keeps the investment name, resolves
    company, and takes the majority CapEx/OpEx across that project's rows. Returns None when it can't be
    computed (no db / no matched Clarity row / no in-period hours) so the caller falls back to one line.
    """
    ts = li.matched_clarity
    if db is None or ts is None or not ts.contractor_name_normalized:
        return None
    start, end = _line_period(li, inv)
    q = select(models.ClarityTimesheet).where(
        models.ClarityTimesheet.contractor_name_normalized == ts.contractor_name_normalized,
        countable_status_filter(),  # Posted OR Submitted
        models.ClarityTimesheet.is_time_off.is_(False),
    )
    if start and end:
        q = q.where(
            models.ClarityTimesheet.date_worked >= start,
            models.ClarityTimesheet.date_worked <= end,
        )
    # project_id -> accumulator
    acc: dict[str | None, dict] = {}
    for e in db.scalars(q).all():
        pid = e.project_id
        a = acc.setdefault(pid, {
            "name": e.investment_name, "hours": 0.0, "capex_votes": defaultdict(float),
        })
        a["hours"] += e.hours or 0.0
        cx = _norm_capex(e.capex_opex)
        if cx:
            a["capex_votes"][cx] += e.hours or 0.0
        if not a["name"] and e.investment_name:
            a["name"] = e.investment_name

    splits: list[ProjectSplit] = []
    for pid, a in acc.items():
        if round(a["hours"], 2) <= 0:
            continue
        votes = a["capex_votes"]
        capex_opex = max(votes, key=votes.get) if votes else None
        splits.append(ProjectSplit(
            project_id=pid,
            project_name=a["name"],
            hours=round(a["hours"], 2),
            company=_company_for_project(pid, a["name"]),
            capex_opex=capex_opex,
        ))
    # Stable, readable order: biggest hours first.
    splits.sort(key=lambda s: s.hours, reverse=True)
    return splits or None


def _line_weight(li: models.InvoiceLineItem) -> float:
    """The line's dollar weight: its invoice amount, else hours x rate, else 0."""
    if li.amount is not None:
        return li.amount
    if li.hours is not None and li.rate is not None:
        return li.hours * li.rate
    return 0.0


def invoice_project_spend(inv: models.Invoice, db: Session | None) -> dict[str, float]:
    """{project_id -> spend} for one invoice: each line's invoice amount allocated across that
    contractor's Clarity projects proportional to hours (so per-project spend sums to the invoice
    total). Same proportional method as the executive report's 'Spend by Project'. A line with no
    project split is credited whole to its matched project id, if any."""
    spend: dict[str, float] = defaultdict(float)
    for li in inv.line_items:
        weight = _line_weight(li)
        if weight == 0:
            continue
        splits = _clarity_project_splits(li, inv, db)
        total = sum(s.hours for s in splits) if splits else 0.0
        if splits and total > 0:
            for s in splits:
                if s.project_id:
                    spend[s.project_id] += weight * (s.hours / total)
        else:
            pid = li.matched_clarity.project_id if li.matched_clarity else None
            if pid:
                spend[pid] += weight
    return {k: round(v, 2) for k, v in spend.items()}


def _line_description(li: models.InvoiceLineItem, inv: models.Invoice) -> str:
    """'Contractor - Project - start-end' (the user-chosen line description shape).

    Project name comes from the matched Clarity timesheet's investment; omitted if unknown.
    Period falls back to the invoice period when the line carries no per-line dates.
    """
    name = li.contractor_name or "Contractor"
    start, end = _line_period(li, inv)
    parts = [name]
    project = li.matched_clarity.investment_name if li.matched_clarity else None
    if project:
        parts.append(project)
    if start and end:
        parts.append(f"{_fmt_date(start)}-{_fmt_date(end)}")
    return " - ".join(parts)  # ASCII separator — keep the whole CSV ASCII-safe for Coupa import


def build_header_row(inv: models.Invoice, *, matched_only: bool = False) -> dict[str, str]:
    """The single `Invoice` header data row, keyed by column name (only populated fields are set;
    every other column defaults to blank when the row is emitted).

    A fully-matched invoice gets the real Chart of Accounts; a draft (flagged invoice, matched lines
    only) leaves it blank since it isn't a clean match. Status is always blank.
    """
    is_clean_match = inv.status == models.STATUS_MATCHED
    return {
        "Invoice": "Invoice",
        "Invoice Number*": inv.invoice_number or "",
        "Supplier Name": inv.vendor_name or "",
        "Supplier Number": supplier_number_for(inv.vendor_name) or "",
        "Status": "",
        "Invoice Date*": _fmt_date(inv.date_received),
        "Submit For Approval?": SUBMIT_FOR_APPROVAL,
        "Line Level Taxation*": LINE_LEVEL_TAXATION,
        # Approver from the copy tracker fills the sample's "<<Approver Name>>" (Requester Name) cell.
        "Requester Name": approver_for(inv.vendor_name, inv.invoice_number) or "",
        "Chart of Accounts*": CHART_OF_ACCOUNTS_MATCHED if is_clean_match else "",
        "Currency": CURRENCY_DEFAULT,
    }


def _line_row(
    inv: models.Invoice, li: models.InvoiceLineItem, line_no: int, *,
    hours: float | None, description: str, segments: dict[str, str],
) -> dict[str, str]:
    """Assemble one `Invoice Line` row. Price* is the invoice rate, Quantity the hours — Coupa
    computes the line total (rate x hours) from them."""
    return {
        "Invoice Line": "Invoice Line",
        "Invoice Number*": inv.invoice_number or "",
        "Supplier Name": inv.vendor_name or "",
        "Supplier Number": supplier_number_for(inv.vendor_name) or "",
        "Line Number": str(line_no),
        "Description*": description,
        "Price*": _fmt_money(li.rate),
        "Quantity": _fmt_num(hours),
        "Unit of Measure*": UNIT_OF_MEASURE,
        **segments,
    }


def _dt_cost_center_map(db: Session | None) -> dict[str, str]:
    """{normalized contractor name -> Clarity department DT code}, for OPEX cost-center coding."""
    if db is None:
        return {}
    from app.services.clarity_resources import cost_center_map  # lazy: avoids import cycle
    return cost_center_map(db)


def _desc_triple(gl_code: str, cost_center: str, project_id: str | None) -> str:
    """The accounting triple appended to a line description: '(GLcode-CostCenter-ProjectID)', e.g.
    '(667040-DT500-PR00099)'. Empty when there's no project id to anchor it."""
    if not project_id:
        return ""
    return f"({gl_code}-{cost_center}-{project_id})"


def _append_triple(description: str, triple: str) -> str:
    return f"{description} {triple}" if triple else description


def build_line_rows(
    inv: models.Invoice, db: Session | None = None, *, matched_only: bool = False
) -> list[dict[str, str]]:
    """`Invoice Line` rows for the invoice.

    One row per (contractor, Clarity project): a contractor's matched in-period hours are split per
    project (Investment), each row carrying that project's hours and its own company code / cost center
    / CapEx-OpEx GL code segments. Cost center is the company code (H0003/AC000) for CapEx work and the
    contractor's Clarity department DT code for OpEx work; the GL code is vendor-aware for OpEx. Each
    description ends with the accounting triple '(GLcode-CostCenter-ProjectID)'.

    A contractor with no Clarity project split (no match / no in-period hours) falls back to one line
    using the invoice's own hours. With `matched_only=True` (draft CSV for a flagged invoice), only
    contractors that matched Clarity get lines; unmatched contractors are skipped to fill in by hand.
    """
    dt_map = _dt_cost_center_map(db)
    rows: list[dict[str, str]] = []
    line_no = 1
    for li in inv.line_items:
        if matched_only and li.line_status != models.STATUS_MATCHED:
            continue
        norm = li.matched_clarity.contractor_name_normalized if li.matched_clarity else None
        dt_code = dt_map.get(norm) if norm else None
        splits = _clarity_project_splits(li, inv, db)

        if splits:
            start, end = _line_period(li, inv)
            period = f"{_fmt_date(start)}-{_fmt_date(end)}" if start and end else ""
            name = li.contractor_name or "Contractor"
            for s in splits:
                cost_center = cost_center_for(s.company, s.capex_opex, dt_code)
                gl_code = gl_code_for(s.capex_opex, inv.vendor_name)
                triple = _desc_triple(gl_code, cost_center, s.project_id)
                base = " - ".join(p for p in (name, s.project_name, period) if p)
                rows.append(
                    _line_row(inv, li, line_no, hours=s.hours,
                              description=_append_triple(base, triple),
                              segments=_segment_values(s.company, cost_center, gl_code))
                )
                line_no += 1
            continue

        # No Clarity split: one line using the invoice's own hours.
        project = _project_for(li, db)
        company = _company_for(li)
        capex_opex = _capex_opex_for(li, project)
        cost_center = cost_center_for(company, capex_opex, dt_code)
        gl_code = gl_code_for(capex_opex, inv.vendor_name)
        pid = li.matched_clarity.project_id if li.matched_clarity else None
        triple = _desc_triple(gl_code, cost_center, pid)
        rows.append(
            _line_row(inv, li, line_no, hours=li.hours,
                      description=_append_triple(_line_description(li, inv), triple),
                      segments=_segment_values(company, cost_center, gl_code))
        )
        line_no += 1
    return rows


def coupa_csv_bytes(
    inv: models.Invoice, db: Session | None = None, *, matched_only: bool = False
) -> bytes:
    """Render the Coupa import CSV for one invoice as UTF-8 bytes (BOM for Excel/Coupa).

    Layout matches documentation/Coupa sample invoice.xlsx: the two record-type schema rows, then the
    `Invoice` header data row, then one `Invoice Line` row per contractor line item. Unpopulated
    columns are emitted blank.

    `matched_only=True` produces a DRAFT for a flagged invoice: only matched contractors get lines.
    """
    buf = io.StringIO(newline="")
    writer = csv.writer(buf)

    # Two record-type schema rows up top, exactly like the sample workbook.
    writer.writerow(INVOICE_HEADER_COLUMNS)
    writer.writerow(INVOICE_LINE_COLUMNS)

    header = build_header_row(inv, matched_only=matched_only)
    writer.writerow([header.get(c, "") for c in INVOICE_HEADER_COLUMNS])
    for row in build_line_rows(inv, db, matched_only=matched_only):
        writer.writerow([row.get(c, "") for c in INVOICE_LINE_COLUMNS])

    return buf.getvalue().encode("utf-8-sig")


def _safe_filename_part(s: str) -> str:
    """Make a string safe to drop into a download filename: strip characters Windows/most OSes
    reject (\\ / : * ? \" < > |), collapse whitespace, and trim trailing dots/spaces."""
    s = re.sub(r'[\\/:*?"<>|]', "", s)
    s = re.sub(r"\s+", " ", s).strip().strip(".")
    return s


def coupa_csv_filename(inv: models.Invoice, *, draft: bool = False) -> str:
    """'<vendor name>_<invoice number>.csv' so each download is easy to tie back to its invoice.

    Draft (flagged, matched-only) downloads are prefixed 'DRAFT_' to keep them distinct.
    """
    vendor = _safe_filename_part(inv.vendor_name or "")
    number = _safe_filename_part(inv.invoice_number or str(inv.id))
    name = "_".join(p for p in (vendor, number) if p) or f"coupa_{inv.id}"
    return f"{'DRAFT_' if draft else ''}{name}.csv"
