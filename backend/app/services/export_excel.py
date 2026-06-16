"""Build an Excel workbook for an invoice: invoice vs. Clarity side-by-side + projects (User Story 19)."""
from __future__ import annotations

from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.schemas import InvoiceDetail

_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_GROUP_FONT = Font(bold=True, size=12)
_BAD_FILL = PatternFill("solid", fgColor="FEE2E2")


def _style_header(ws, row: int, ncols: int) -> None:
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _autosize(ws) -> None:
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(width + 2, 10), 50)


def workbook_from_detail(d: InvoiceDetail) -> BytesIO:
    wb = Workbook()

    # --- Summary ---
    ws = wb.active
    ws.title = "Summary"
    rows = [
        ("Vendor", d.vendor_name),
        ("Invoice #", d.invoice_number),
        ("Date received", str(d.date_received) if d.date_received else ""),
        ("Period", f"{d.payment_period_start or ''} - {d.payment_period_end or ''}"),
        ("Total invoice cost", d.total_invoice_cost),
        ("Status", d.status),
        ("Routed to", d.routed_to or ""),
        ("Lines matched", f"{d.matched_line_count}/{d.line_item_count}"),
    ]
    ws["A1"] = "Invoice Summary"
    ws["A1"].font = _GROUP_FONT
    for i, (k, v) in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=i, column=2, value=v)
    r = len(rows) + 3
    if d.mismatch_reasons:
        ws.cell(row=r, column=1, value="What didn't match").font = _GROUP_FONT
        for m in d.mismatch_reasons:
            r += 1
            ws.cell(row=r, column=1, value=m.field)
            ws.cell(row=r, column=2, value=m.reason)
    _autosize(ws)

    # --- Comparison (Invoice vs Clarity side-by-side) ---
    cmp = wb.create_sheet("Invoice vs Clarity")
    cmp.cell(row=1, column=1, value="INVOICE DATA").font = _GROUP_FONT
    cmp.cell(row=1, column=6, value="CLARITY DATA").font = _GROUP_FONT
    headers = ["Contractor", "Hours", "Rate", "Amount", "", "Contractor", "Hours", "Δ Hours", "Match", "Status"]
    for c, h in enumerate(headers, start=1):
        cmp.cell(row=2, column=c, value=h)
    _style_header(cmp, 2, len(headers))

    row = 3
    for li in d.line_items:
        diff = li.diff or {}
        vals = [
            li.contractor_name, li.hours, li.rate, li.amount, "",
            diff.get("clarity_name"), diff.get("clarity_hours"), diff.get("hours_delta"),
            diff.get("match_method"), li.line_status,
        ]
        for c, v in enumerate(vals, start=1):
            cmp.cell(row=row, column=c, value=v)
        if li.line_status and li.line_status != "matched":
            for c in range(1, len(headers) + 1):
                cmp.cell(row=row, column=c).fill = _BAD_FILL
        row += 1
    _autosize(cmp)

    # --- Clarity timesheets (aggregated per contractor) ---
    ts = wb.create_sheet("Clarity Timesheets")
    th = ["Contractor", "Hours", "Project", "Investment Manager"]
    for c, h in enumerate(th, start=1):
        ts.cell(row=1, column=c, value=h)
    _style_header(ts, 1, len(th))
    for i, t in enumerate(d.clarity_timesheets, start=2):
        ts.cell(row=i, column=1, value=t.contractor_name)
        ts.cell(row=i, column=2, value=t.hours)
        ts.cell(row=i, column=3, value=t.project_id)
        ts.cell(row=i, column=4, value=t.investment_manager)
    _autosize(ts)

    # --- Projects ---
    pj = wb.create_sheet("Projects")
    ph = ["Type", "Project ID", "Project", "Budget ID", "Cost Center", "Vendor", "LOB", "Spend"]
    for c, h in enumerate(ph, start=1):
        pj.cell(row=1, column=c, value=h)
    _style_header(pj, 1, len(ph))
    for i, p in enumerate(d.clarity_projects, start=2):
        for c, v in enumerate(
            [p.capex_opex, p.project_id, p.project_name, p.budget_id, p.cost_center, p.vendor, p.lob, p.spend],
            start=1,
        ):
            pj.cell(row=i, column=c, value=v)
    _autosize(pj)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
